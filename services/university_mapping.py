from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz


class UniversityMappingService:
    """
    Deterministic university -> CAP institute mapping.

    Matching policy:
    1. Institute Code is the primary key.
    2. Name matching is used only for mapping rows missing Institute Code.
    3. Fallback name matching is strict and location-aware.
    """

    def __init__(self, mapping_path: str | Path | None = None) -> None:
        project_root = Path(__file__).resolve().parent.parent
        self.mapping_path = Path(
            mapping_path
            or project_root / "data" / "university_college_mapping.xlsx"
        )
        self._rows: list[dict[str, str]] | None = None

    def list_universities(self) -> list[str]:
        output: list[str] = []
        seen: set[str] = set()

        for row in self._load_rows():
            university = row.get("University", "").strip()
            key = self._norm(university)
            if university and key not in seen:
                seen.add(key)
                output.append(university)

        return output

    def resolve(self, requested_university: str | None) -> dict[str, Any]:
        requested = str(requested_university or "").strip()

        if not requested:
            return {
                "requested": False,
                "requested_university": None,
                "matched_university": None,
                "university_code": None,
                "colleges": [],
                "institute_codes": [],
                "rows": [],
                "status": "no_university_requested",
            }

        rows = self._load_rows()
        groups: dict[str, list[dict[str, str]]] = {}
        names: dict[str, str] = {}
        codes: dict[str, str] = {}

        for row in rows:
            university = row.get("University", "").strip()
            if not university:
                continue

            key = self._norm(university)
            groups.setdefault(key, []).append(row)
            names[key] = university
            codes[key] = row.get("University Code", "").strip()

        requested_norm = self._norm(requested)
        requested_compact = self._compact(requested)
        matched_key: str | None = None

        if requested_norm in groups:
            matched_key = requested_norm

        if matched_key is None:
            for key, code in codes.items():
                if requested_compact and requested_compact == self._compact(code):
                    matched_key = key
                    break

        if matched_key is None:
            candidates = [
                key
                for key in groups
                if requested_norm
                and (requested_norm in key or key in requested_norm)
            ]
            if len(candidates) == 1:
                matched_key = candidates[0]

        if matched_key is None:
            scored = sorted(
                (
                    (fuzz.WRatio(requested_norm, key), key)
                    for key in groups
                ),
                reverse=True,
            )
            if scored and scored[0][0] >= 88:
                if len(scored) == 1 or scored[0][0] - scored[1][0] >= 5:
                    matched_key = scored[0][1]

        if matched_key is None:
            return {
                "requested": True,
                "requested_university": requested,
                "matched_university": None,
                "university_code": None,
                "colleges": [],
                "institute_codes": [],
                "rows": [],
                "available_universities": self.list_universities(),
                "status": "university_not_found",
            }

        matched_rows = groups[matched_key]
        colleges: list[str] = []
        institute_codes: list[str] = []
        seen_colleges: set[str] = set()
        seen_codes: set[str] = set()

        for row in matched_rows:
            college = row.get("College", "").strip()
            college_key = self._norm(college)
            if college and college_key not in seen_colleges:
                seen_colleges.add(college_key)
                colleges.append(college)

            institute_code = self._normalise_institute_code(
                row.get("Institute Code")
            )
            if institute_code and institute_code not in seen_codes:
                seen_codes.add(institute_code)
                institute_codes.append(institute_code)

        return {
            "requested": True,
            "requested_university": requested,
            "matched_university": names[matched_key],
            "university_code": codes.get(matched_key) or None,
            "colleges": colleges,
            "institute_codes": institute_codes,
            "rows": matched_rows,
            "status": "resolved",
        }

    def filter_cutoff_records(
        self,
        records: list[dict[str, Any]],
        requested_university: str | None,
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        resolution = self.resolve(requested_university)

        report: dict[str, Any] = {
            "requested_university": resolution.get("requested_university"),
            "matched_university": resolution.get("matched_university"),
            "mapping_status": resolution.get("status"),
            "mapping_college_count": len(resolution.get("colleges") or []),
            "mapping_institute_code_count": len(
                resolution.get("institute_codes") or []
            ),
            "before": len(records),
            "after": len(records),
            "matched_by_institute_code": 0,
            "matched_by_fallback_name": 0,
            "matched_cap_colleges": [],
        }

        if not resolution.get("requested"):
            return records, report

        if resolution.get("status") != "resolved":
            report["after"] = 0
            return [], report

        allowed_codes = {
            self._normalise_institute_code(code)
            for code in (resolution.get("institute_codes") or [])
            if self._normalise_institute_code(code)
        }

        # Only rows without Institute Code may use fallback matching.
        fallback_rows = [
            {
                "college_norm": self._norm(row.get("College")),
                "college_core": self._college_core(row.get("College")),
                "city": self._norm(row.get("City")),
                "district": self._norm(row.get("District")),
            }
            for row in (resolution.get("rows") or [])
            if str(row.get("College") or "").strip()
            and not self._normalise_institute_code(row.get("Institute Code"))
        ]

        filtered: list[dict[str, Any]] = []
        matched_names: list[str] = []
        seen_names: set[str] = set()
        code_match_count = 0
        fallback_match_count = 0

        for record in records:
            metadata = record.get("metadata") or {}

            cap_code = self._normalise_institute_code(
                metadata.get("institute_code")
                or metadata.get("institute")
                or metadata.get("college_code")
            )

            matched = False
            match_method = ""

            # Primary exact Institute Code match.
            if cap_code and cap_code in allowed_codes:
                matched = True
                match_method = "institute_code"

            # Strict fallback only for mapping rows missing a code.
            if not matched and fallback_rows:
                college = str(
                    metadata.get("college")
                    or metadata.get("institute_name")
                    or ""
                ).strip()

                location_text = " ".join(
                    str(metadata.get(key) or "")
                    for key in ("location", "city", "district")
                )

                if college and self._matches_fallback_row(
                    cap_college=college,
                    cap_location=location_text,
                    mapped_rows=fallback_rows,
                ):
                    matched = True
                    match_method = "fallback_name"

            if not matched:
                continue

            filtered.append(record)

            if match_method == "institute_code":
                code_match_count += 1
            else:
                fallback_match_count += 1

            college_name = str(
                metadata.get("college")
                or metadata.get("institute_name")
                or ""
            ).strip()

            college_key = self._norm(college_name)
            if college_name and college_key not in seen_names:
                seen_names.add(college_key)
                matched_names.append(college_name)

        report["after"] = len(filtered)
        report["matched_by_institute_code"] = code_match_count
        report["matched_by_fallback_name"] = fallback_match_count
        report["matched_cap_colleges"] = matched_names

        return filtered, report

    def _matches_fallback_row(
        self,
        cap_college: str,
        cap_location: str,
        mapped_rows: list[dict[str, str]],
    ) -> bool:
        cap_norm = self._norm(cap_college)
        cap_core = self._college_core(cap_college)
        cap_context = self._norm(f"{cap_college} {cap_location}")

        for row in mapped_rows:
            mapped_norm = row["college_norm"]
            mapped_core = row["college_core"]

            if not mapped_norm or not mapped_core:
                continue

            exact_name = cap_norm == mapped_norm
            very_strong_fuzzy = fuzz.WRatio(cap_core, mapped_core) >= 96

            if not (exact_name or very_strong_fuzzy):
                continue

            city = row["city"]
            district = row["district"]

            # If mapping has location, it must agree.
            if city or district:
                if city and city in cap_context:
                    return True
                if district and district in cap_context:
                    return True
                continue

            return True

        return False

    def _load_rows(self) -> list[dict[str, str]]:
        if self._rows is not None:
            return self._rows

        if not self.mapping_path.exists():
            raise FileNotFoundError(
                f"University mapping file not found: {self.mapping_path}"
            )

        workbook = load_workbook(
            self.mapping_path,
            read_only=True,
            data_only=True,
        )

        try:
            sheet = (
                workbook["University College Mapping"]
                if "University College Mapping" in workbook.sheetnames
                else workbook[workbook.sheetnames[0]]
            )

            iterator = sheet.iter_rows(values_only=True)
            header_values = next(iterator, None)

            if not header_values:
                self._rows = []
                return self._rows

            headers = [
                str(value or "").strip()
                for value in header_values
            ]

            required = {
                "University Code",
                "University",
                "College",
                "Institute Code",
                "City",
                "District",
            }

            missing = required.difference(headers)
            if missing:
                raise ValueError(
                    "University mapping missing required columns: "
                    + ", ".join(sorted(missing))
                )

            rows: list[dict[str, str]] = []

            for values in iterator:
                row = {
                    header: str(value or "").strip()
                    for header, value in zip(headers, values)
                }

                if row.get("University") and row.get("College"):
                    rows.append(row)

            self._rows = rows
            return rows
        finally:
            workbook.close()

    @staticmethod
    def _normalise_institute_code(value: Any) -> str:
        text = str(value or "").strip()

        if text.endswith(".0"):
            text = text[:-2]

        return re.sub(
            r"[^A-Z0-9]",
            "",
            text.upper(),
        )

    @staticmethod
    def _norm(value: Any) -> str:
        text = str(value or "").casefold()
        text = re.sub(r"[&/]+", " ", text)
        text = re.sub(r"[^a-z0-9]+", " ", text)
        return " ".join(text.split())

    @staticmethod
    def _compact(value: Any) -> str:
        return re.sub(
            r"[^a-z0-9]",
            "",
            str(value or "").casefold(),
        )

    @classmethod
    def _college_core(cls, value: Any) -> str:
        text = re.sub(
            r"\([^)]*\)",
            " ",
            str(value or ""),
        )
        return cls._norm(text)


def get_university_mapping_service() -> UniversityMappingService:
    return UniversityMappingService()