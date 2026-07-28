from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelService:
    def export_recommendations(
        self,
        recommendations: list[dict[str, Any]],
        output_path: str | Path,
    ) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        rows = []
        for item in recommendations:
            rows.append({
                "Rank": item.get("rank"),
                "Zone": item.get("zone"),
                "College": item.get("college"),
                "Branch": item.get("branch"),
                "Location": item.get("location"),
                "Seat Type": item.get("category_or_seat_type"),
                "Seat Allocation": item.get("seat_allocation"),
                "Historical Cutoff": item.get("historical_cutoff"),
                "Student Percentile": item.get("student_percentile"),
                "Cutoff Gap": item.get("cutoff_gap"),
                "Reason": item.get("reason"),
                "Evidence IDs": ", ".join(map(str, item.get("evidence_ids", []))),
            })

        pd.DataFrame(rows).to_excel(path, index=False, sheet_name="Recommendations")
        self._format(path)
        return path

    @staticmethod
    def _format(path: Path) -> None:
        workbook = load_workbook(path)
        sheet = workbook["Recommendations"]
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)

        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            width = max(
                len(str(sheet.cell(row=row, column=column).value or ""))
                for row in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[letter].width = min(max(width + 2, 12), 55)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(path)
