from __future__ import annotations

import base64
from pathlib import Path
from typing import Any

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook

# ==========================================================
# Configuration
# ==========================================================

API_URL = "http://127.0.0.1:8000"

LOGO_PATHS = (
    Path(__file__).with_name("logo_2.jpeg"),
    Path(__file__).with_name("logo_2.jpg"),
    Path(__file__).with_name("logo_2.png"),
    Path(__file__).with_name("logo_2(3).jpeg"),
)

CATEGORY_OPTIONS = [
    "OPEN", "OBC", "SC", "ST", "EWS", "NT-A", "NT-B", "NT-C", "NT-D", "VJ", "SBC"
]

BRANCH_OPTIONS = [
    "Any",
    "Computer Engineering",
    "Computer Science and Engineering",
    "Information Technology",
    "Artificial Intelligence",
    "Artificial Intelligence and Data Science",
    "Data Science",
    "Cyber Security",
    "Computer Science and Business Systems",
    "Electronics and Telecommunication Engineering",
    "Electronics Engineering",
    "Electrical Engineering",
    "Mechanical Engineering",
    "Civil Engineering",
    "Chemical Engineering",
    "Instrumentation Engineering",
    "Robotics and Automation",
]

SEAT_TYPE_OPTIONS = [
    "Any", "Government", "Government Aided", "Autonomous", "Private", "Minority"
]

LOCATION_OPTIONS = [
    "Any", "Pune", "Mumbai", "Navi Mumbai", "Thane", "Nagpur", "Nashik",
    "Kolhapur", "Sangli", "Satara", "Chhatrapati Sambhajinagar", "Amravati",
    "Solapur", "Ahmednagar",
]

GENDER_OPTIONS = [
    "Not Specified", "Male", "Female", "Other", "Prefer Not To Say"
]


COLLEGE_COUNT_OPTIONS = [10, 20, 30, 40, 50]


UNIVERSITY_MAPPING_PATH = Path(__file__).parent / "data" / "university_college_mapping.xlsx"


def _load_university_options() -> list[str]:
    if not UNIVERSITY_MAPPING_PATH.exists():
        return ["Any"]

    wb = load_workbook(UNIVERSITY_MAPPING_PATH, read_only=True, data_only=True)
    try:
        ws = (
            wb["University College Mapping"]
            if "University College Mapping" in wb.sheetnames
            else wb[wb.sheetnames[0]]
        )
        rows = ws.iter_rows(values_only=True)
        headers = [str(v or "").strip() for v in (next(rows, None) or [])]
        if "University" not in headers:
            return ["Any"]

        index = headers.index("University")
        output = ["Any"]
        seen: set[str] = set()

        for row in rows:
            if index >= len(row):
                continue
            name = str(row[index] or "").strip()
            key = name.casefold()
            if name and key not in seen:
                seen.add(key)
                output.append(name)

        return output
    finally:
        wb.close()


UNIVERSITY_OPTIONS = _load_university_options()



# ==========================================================
# Helpers
# ==========================================================


def _without_any(values: list[str]) -> list[str]:
    return [] if "Any" in values else values


def _logo_data_uri() -> str | None:
    for path in LOGO_PATHS:
        if path.exists():
            suffix = path.suffix.lower()
            mime = "image/png" if suffix == ".png" else "image/jpeg"
            encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
            return f"data:{mime};base64,{encoded}"
    return None


def _error_message(response: requests.Response) -> str:
    try:
        data = response.json()
        detail = data.get("detail", data) if isinstance(data, dict) else data
        return str(detail)
    except ValueError:
        return response.text or f"HTTP {response.status_code}"


def _post_json(endpoint: str, body: dict[str, Any], timeout: int) -> dict[str, Any]:
    response = requests.post(f"{API_URL}{endpoint}", json=body, timeout=timeout)
    if not response.ok:
        raise RuntimeError(_error_message(response))
    return response.json()


def _normalise_zone(value: Any) -> str:
    text = str(value or "").strip().casefold()
    if text in {"dream", "aspirational", "reach"}:
        return "Dream"
    if text in {"target", "competitive", "match"}:
        return "Target"
    if text in {"safe", "safer", "realistic"}:
        return "Safer"
    return str(value or "Other").title()


def _zone_counts(recommendations: list[dict[str, Any]]) -> dict[str, int]:
    counts = {"Dream": 0, "Target": 0, "Safer": 0}
    for item in recommendations:
        zone = _normalise_zone(
            item.get("zone")
            or item.get("recommendation_band")
            or item.get("band")
        )
        if zone in counts:
            counts[zone] += 1
    return counts


def _prepare_table(recommendations: list[dict[str, Any]]) -> pd.DataFrame:
    frame = pd.DataFrame(recommendations)
    if frame.empty:
        return frame

    for zone_column in ("zone", "recommendation_band", "band"):
        if zone_column in frame.columns:
            frame[zone_column] = frame[zone_column].map(_normalise_zone)

    preferred_columns = [
        "rank",
        "college",
        "college_name",
        "branch",
        "seat_type",
        "category_or_seat_type",
        "category",
        "historical_cutoff",
        "cutoff_percentile",
        "student_percentile",
        "your_percentile",
        "cutoff_gap",
        "gap",
        "zone",
        "recommendation_band",
        "band",
        "reason",
        "location",
    ]
    visible = [column for column in preferred_columns if column in frame.columns]
    remaining = [
        column
        for column in frame.columns
        if column not in visible and column != "evidence_ids"
    ]
    frame = frame[visible + remaining]
    frame.columns = [column.replace("_", " ").title() for column in frame.columns]
    return frame


def _assistant_response_text(result: dict[str, Any]) -> str:
    """Extract a concise reply; structured pending details are rendered separately."""
    feedback_ui = result.get("feedback_ui")
    if isinstance(feedback_ui, dict) and feedback_ui.get("type") == "pending_choice":
        message = str(feedback_ui.get("message") or "").strip()
        if message:
            return message
        return "I found the requested option and prepared safer alternatives. Choose an action below."

    responses = result.get("counsellor_responses")
    if isinstance(responses, list):
        combined = "\n\n".join(
            str(item).strip() for item in responses if str(item).strip()
        )
        if combined:
            return combined

    for key in ("counsellor_response", "response", "message", "summary"):
        value = str(result.get(key) or "").strip()
        if value:
            return value

    return "I reviewed your request and updated the counselling result."


def _append_chat_message(role: str, content: str) -> None:
    text = str(content or "").strip()
    if text:
        st.session_state.chat_history.append({"role": role, "content": text})


def _format_number(value: Any, digits: int = 2) -> str:
    try:
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return "—"


def _candidate_title(candidate: dict[str, Any]) -> str:
    college = str(candidate.get("college") or "College").strip()
    branch = str(candidate.get("branch") or "Branch").strip()
    return f"{college} — {branch}"


def _send_feedback_action(
    session_id: str,
    feedback: str,
    visible_message: str,
) -> None:
    """Send a typed or button-generated instruction through the same feedback API."""
    _append_chat_message("user", visible_message)
    try:
        with st.spinner("Feedback Agent is analysing your choice..."):
            updated_result = _post_json(
                "/feedback",
                {"session_id": session_id, "feedback": feedback},
                1200,
            )
        st.session_state.result = updated_result
        _append_chat_message("assistant", _assistant_response_text(updated_result))
    except (requests.RequestException, RuntimeError) as exc:
        _append_chat_message("assistant", f"I could not process that request: {exc}")
    st.session_state.chat_input_version += 1
    st.rerun()


def _render_candidate_card(candidate: dict[str, Any], accent: str = "") -> None:
    gap = candidate.get("gap")
    try:
        gap_text = f"{float(gap):+.2f}"
    except (TypeError, ValueError):
        gap_text = "—"
    risk = str(candidate.get("risk_level") or "Not stated")
    st.markdown(
        f"""
<div class="choice-card {accent}">
  <div class="choice-title">{_candidate_title(candidate)}</div>
  <div class="choice-metrics">
    <span><b>Cutoff:</b> {_format_number(candidate.get('historical_cutoff'))}</span>
    <span><b>Your percentile:</b> {_format_number(candidate.get('student_percentile'))}</span>
    <span><b>Gap:</b> {gap_text}</span>
    <span><b>Risk:</b> {risk}</span>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )


def _render_feedback_actions(result: dict[str, Any], session_id: str | None) -> None:
    """Render a compact pending-decision panel with collapsed alternatives."""
    if not session_id:
        return

    ui = result.get("feedback_ui")
    pending = result.get("pending_feedback_action")
    if not isinstance(ui, dict) or ui.get("type") != "pending_choice" or not pending:
        return

    requested = [
        item for item in ui.get("requested_options", [])
        if isinstance(item, dict)
    ]
    same_college = [
        item for item in ui.get("same_college_alternatives", [])
        if isinstance(item, dict)
    ]
    same_branch = [
        item for item in ui.get("same_branch_alternatives", [])
        if isinstance(item, dict)
    ]

    # Compact admission-analysis card.
    st.markdown("#### ⚠️ Admission Analysis")
    for candidate in requested[:1]:
        _render_candidate_card(candidate, "requested-card")

    reason = str(ui.get("confirmation_reason") or "").strip()
    if reason:
        st.caption(reason)

    # Alternatives remain collapsed until the student chooses to inspect them.
    if same_college:
        with st.expander(
            f"🎓 Better Branches ({len(same_college)})",
            expanded=False,
        ):
            for index, candidate in enumerate(same_college[:8]):
                row_left, row_right = st.columns([7.5, 1.25], vertical_alignment="center")
                with row_left:
                    _render_candidate_card(candidate, "compact-alternative-card")
                with row_right:
                    if st.button(
                        "➕ Add",
                        key=f"choose_same_college_{candidate.get('candidate_id')}_{index}",
                        use_container_width=True,
                    ):
                        _send_feedback_action(
                            session_id,
                            f"UI_ACTION::CHOOSE_SAME_COLLEGE::{candidate.get('candidate_id')}",
                            f"Add {_candidate_title(candidate)}",
                        )

    if same_branch:
        with st.expander(
            f"🏫 Better Colleges ({len(same_branch)})",
            expanded=False,
        ):
            for index, candidate in enumerate(same_branch[:8]):
                row_left, row_right = st.columns([7.5, 1.25], vertical_alignment="center")
                with row_left:
                    _render_candidate_card(candidate, "compact-alternative-card")
                with row_right:
                    if st.button(
                        "➕ Add",
                        key=f"choose_same_branch_{candidate.get('candidate_id')}_{index}",
                        use_container_width=True,
                    ):
                        _send_feedback_action(
                            session_id,
                            f"UI_ACTION::CHOOSE_SAME_BRANCH::{candidate.get('candidate_id')}",
                            f"Add {_candidate_title(candidate)}",
                        )

    # Compact single-row action toolbar.
    actions = [item for item in ui.get("actions", []) if isinstance(item, dict)]
    if actions:
        st.markdown('<div class="compact-action-title">Choose an action</div>', unsafe_allow_html=True)
        action_columns = st.columns(len(actions), gap="small", vertical_alignment="center")
        for index, action in enumerate(actions):
            with action_columns[index]:
                button_type = "primary" if action.get("kind") == "primary" else "secondary"
                if st.button(
                    str(action.get("label") or "Continue"),
                    key=f"pending_action_{action.get('id')}_{index}",
                    type=button_type,
                    use_container_width=True,
                ):
                    _send_feedback_action(
                        session_id,
                        str(action.get("feedback") or ""),
                        str(action.get("label") or "Continue"),
                    )



# ==========================================================
# Page and state
# ==========================================================

st.set_page_config(
    page_title="MHT-CET Agentic RAG Counsellor V2",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="collapsed",
)

DEFAULT_STATE: dict[str, Any] = {
    "show_candidate_panel": True,
    "show_knowledge_panel": True,
    "result": None,
    "upload_result": None,
    "chat_history": [],
    "chat_input_version": 0,
}
for state_key, state_value in DEFAULT_STATE.items():
    if state_key not in st.session_state:
        st.session_state[state_key] = state_value


# ==========================================================
# Styling
# ==========================================================

st.markdown(
    """
<style>
:root {
    --navy: #101f42;
    --muted: #667085;
    --line: #e4e8f0;
    --panel: #f7f9fd;
    --dream: #1688ff;
    --target: #ff7900;
    --safer: #15a34a;
    --action: #ff363d;
}

html, body, [class*="css"] {
    font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
}

[data-testid="stSidebar"], [data-testid="collapsedControl"] { display: none !important; }
header[data-testid="stHeader"] {
    display: none !important;
    height: 0 !important;
}
#MainMenu, footer,
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="stStatusWidget"],
.stDeployButton {
    display: none !important;
    visibility: hidden !important;
}

html, body, [data-testid="stAppViewContainer"], .stApp {
    min-height: 100vh !important;
    background: #f8faff !important;
}

.block-container {
    max-width: 100% !important;
    padding: 0.3rem 0.55rem 0.75rem !important;
}

[data-testid="stHorizontalBlock"] {
    gap: 0.65rem !important;
    align-items: stretch !important;
}

[data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
    display: flex !important;
    flex-direction: column !important;
    min-height: auto !important;
}

[data-testid="stHorizontalBlock"] > [data-testid="stColumn"] > div {
    flex: 1 1 auto !important;
}

[data-testid="stVerticalBlock"] { gap: 0.48rem !important; }

/* Make the three main cards use the available viewport height. */
[data-testid="stHorizontalBlock"] > [data-testid="stColumn"] > div > [data-testid="stVerticalBlockBorderWrapper"] {
    height: 100% !important;
    min-height: calc(100vh - 1rem) !important;
}

/* Native Streamlit bordered containers used as our panels */
[data-testid="stVerticalBlockBorderWrapper"] {
    border: 1px solid var(--line) !important;
    border-radius: 10px !important;
    background: #ffffff !important;
    box-shadow: 0 4px 18px rgba(16, 31, 66, 0.045) !important;
    overflow: hidden !important;
}

[data-testid="stVerticalBlockBorderWrapper"] > div {
    height: 100% !important;
}

.side-heading {
    color: var(--navy);
    font-weight: 850;
    font-size: 0.94rem;
    margin: 0.05rem 0 0.35rem;
}

.side-note {
    color: var(--muted);
    font-size: 0.72rem;
    line-height: 1.35;
    margin-top: 0.25rem;
}

.hero {
    position: relative;
    min-height: 126px;
    padding: 0.45rem 0.75rem 0.65rem;
    border-bottom: 1px solid var(--line);
}

.hero-logo {
    position: absolute;
    top: 0.2rem;
    left: 0.05rem;
    width: 154px;
    max-height: 72px;
    object-fit: contain;
}

.hero-title {
    color: var(--navy);
    text-align: center;
    font-size: 1.88rem;
    line-height: 1.18;
    font-weight: 900;
    letter-spacing: -0.035em;
    margin: 1.18rem 9.2rem 0.48rem;
}

.hero-subtitle {
    color: var(--muted);
    text-align: center;
    font-size: 0.88rem;
    margin: 0 4rem;
}

.dashboard-title {
    color: var(--navy);
    font-weight: 850;
    font-size: 1.03rem;
    margin: 0.75rem 0 0.45rem;
}

.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 0.65rem;
    margin-bottom: 0.65rem;
}

.kpi-card {
    border: 1px solid var(--line);
    border-radius: 9px;
    background: #fff;
    padding: 0.68rem 0.78rem;
    box-shadow: 0 3px 12px rgba(16, 31, 66, 0.045);
}

.kpi-label { color: var(--muted); font-size: 0.72rem; margin-bottom: 0.18rem; }
.kpi-value { color: var(--navy); font-size: 1.35rem; font-weight: 900; line-height: 1; }
.kpi-value.dream { color: var(--dream); }
.kpi-value.target { color: var(--target); }
.kpi-value.safer { color: var(--safer); }

.empty-state {
    border: 1px dashed #d3dae7;
    border-radius: 12px;
    text-align: center;
    padding: 3.9rem 1.5rem;
    color: var(--muted);
    background: #fff;
}

.empty-title {
    color: var(--navy);
    font-weight: 850;
    font-size: 1.18rem;
    margin: 0.5rem 0 0.32rem;
}

.status-ready {
    border: 1px solid #d4f2df;
    border-radius: 8px;
    background: #ecfaf2;
    color: #168447;
    font-size: 0.78rem;
    padding: 0.63rem 0.76rem;
    margin-top: 0.58rem;
}

.stButton > button {
    min-height: 2.25rem !important;
    border-radius: 7px !important;
    border: 1px solid #d7dce6 !important;
    font-weight: 700 !important;
}

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #ff3038, #ff4a50) !important;
    border: none !important;
    color: #fff !important;
}

[data-testid="stFileUploaderDropzone"] {
    border: 1px dashed #b8c4df !important;
    border-radius: 8px !important;
    background: #fff !important;
    padding: 0.7rem 0.4rem !important;
}

[data-testid="stFileUploaderDropzoneInstructions"] span { font-size: 0.74rem !important; }
label[data-testid="stWidgetLabel"] p { color: #344054; font-size: 0.73rem; font-weight: 650; }
[data-baseweb="select"] > div, [data-testid="stNumberInput"] input, textarea { border-radius: 7px !important; }
[data-testid="stDataFrame"] { border: 1px solid #e8ecf3; border-radius: 7px; overflow: hidden; }


.chat-panel {
    border: 1px solid #e1e7f0;
    border-radius: 12px;
    background: #f8faff;
    padding: 0.72rem 0.78rem 0.35rem;
    margin-top: 0.72rem;
}

.chat-heading {
    color: var(--navy);
    font-size: 1rem;
    font-weight: 850;
    margin-bottom: 0.45rem;
}

[data-testid="stChatMessage"] {
    border: 1px solid #e4e9f2;
    border-radius: 12px;
    background: #ffffff;
    padding: 0.32rem 0.48rem;
    margin-bottom: 0.5rem;
}

[data-testid="stChatInput"] {
    margin-top: 0.45rem;
}

.quick-action-note {
    color: var(--muted);
    font-size: 0.72rem;
    margin: 0.2rem 0 0.4rem;
}


.decision-panel {
    border: 1px solid #dfe6f1;
    border-radius: 12px;
    background: #fbfcff;
    padding: 0.8rem;
    margin: 0.55rem 0 0.75rem;
}

.choice-card {
    border: 1px solid #e1e7f0;
    border-radius: 10px;
    background: #ffffff;
    padding: 0.7rem 0.8rem;
    margin: 0.3rem 0;
}

.requested-card {
    border-left: 4px solid #e5484d;
    background: #fffafa;
}

.alternative-card {
    border-left: 4px solid #2f80ed;
}

.choice-title {
    color: var(--navy);
    font-weight: 800;
    font-size: 0.88rem;
    margin-bottom: 0.38rem;
}

.choice-metrics {
    display: flex;
    flex-wrap: wrap;
    gap: 0.35rem 0.9rem;
    color: #475467;
    font-size: 0.76rem;
}


/* Compact feedback decision controls */
[data-testid="stExpander"] {
    border: 1px solid #e1e7f0 !important;
    border-radius: 9px !important;
    background: #ffffff !important;
    margin: 0.35rem 0 !important;
}

[data-testid="stExpander"] summary {
    min-height: 2.35rem !important;
    padding: 0.25rem 0.7rem !important;
    font-weight: 750 !important;
    color: var(--navy) !important;
}

[data-testid="stExpanderDetails"] {
    padding: 0.15rem 0.55rem 0.45rem !important;
}

.compact-alternative-card {
    padding: 0.48rem 0.62rem !important;
    margin: 0.18rem 0 !important;
}

.compact-alternative-card .choice-title {
    font-size: 0.8rem !important;
    margin-bottom: 0.2rem !important;
}

.compact-alternative-card .choice-metrics {
    font-size: 0.69rem !important;
    gap: 0.18rem 0.65rem !important;
}

.compact-action-title {
    color: var(--navy);
    font-weight: 800;
    font-size: 0.9rem;
    margin: 0.55rem 0 0.25rem;
}

/* Keep nested feedback rows and button toolbars content-height only. */
[data-testid="stExpander"] [data-testid="stColumn"],
[data-testid="stExpander"] [data-testid="stHorizontalBlock"],
[data-testid="stVerticalBlock"] [data-testid="stHorizontalBlock"] [data-testid="stColumn"] {
    min-height: 0 !important;
    height: auto !important;
}

[data-testid="stExpander"] .stButton > button,
.compact-action-title + div .stButton > button {
    min-height: 2.05rem !important;
    padding: 0.25rem 0.55rem !important;
    white-space: nowrap !important;
    font-size: 0.78rem !important;
}

@media (max-width: 1150px) {
    [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
        min-height: auto !important;
    }
    [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] > div > [data-testid="stVerticalBlockBorderWrapper"] {
        min-height: auto !important;
    }
    .hero-logo { position: static; display: block; margin: 0 auto; width: 135px; }
    .hero-title { margin: 0.35rem 0.5rem; font-size: 1.5rem; }
    .hero-subtitle { margin: 0; }
    .kpi-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
}
</style>
""",
    unsafe_allow_html=True,
)


# ==========================================================
# Responsive three-column workspace
# ==========================================================

show_left = bool(st.session_state.show_candidate_panel)
show_right = bool(st.session_state.show_knowledge_panel)

if show_left and show_right:
    column_widths = [1.55, 5.35, 1.55]
elif show_left:
    column_widths = [1.62, 6.7, 0.28]
elif show_right:
    column_widths = [0.28, 6.7, 1.62]
else:
    column_widths = [0.28, 8.1, 0.28]

left_column, center_column, right_column = st.columns(column_widths, gap="small")


# ==========================================================
# Candidate profile panel
# ==========================================================

with left_column:
    if show_left:
        with st.container(border=True):
            if st.button(
                "☰  Candidate Profile",
                key="candidate_panel_toggle_close",
                use_container_width=True,
                help="Hide Candidate Profile",
            ):
                st.session_state.show_candidate_panel = False
                st.rerun()

            st.markdown('<div class="side-heading">📋 Candidate Profile Controls</div>', unsafe_allow_html=True)

            percentile = st.number_input(
                "Percentile",
                min_value=0.0,
                max_value=100.0,
                value=90.0,
                step=0.01,
                key="percentile",
            )
            category = st.selectbox("Category", CATEGORY_OPTIONS, key="category")
            gender = st.selectbox("Gender", GENDER_OPTIONS, key="gender")
            branches = st.multiselect(
                "Preferred branches",
                BRANCH_OPTIONS,
                default=["Computer Engineering", "Information Technology"],
                help="Select Any to remove the branch restriction.",
                key="branches",
            )
            location_text = st.text_input(
                "Preferred location(s)",
                placeholder="Example: Pune, Mumbai, Nagpur",
                help="Enter one or more cities separated by commas. Leave blank for all locations.",
                key="location_text",
            )

            locations = [
                city.strip()
                for city in location_text.split(",")
                if city.strip()
            ]
            
            preferred_university = st.selectbox(
                "Preferred University",
                UNIVERSITY_OPTIONS,
                index=0,
                key="preferred_university",
                help=(
                    "Select Any for statewide recommendations. "
                    "A selected university is applied as a hard filter before ranking."
                ),
            )

            seat_type = st.selectbox("Seat type", SEAT_TYPE_OPTIONS, key="seat_type")
        
            college_count = st.selectbox(
                "College count",
                COLLEGE_COUNT_OPTIONS,
                index=2,
                key="college_count",
            )
            user_request = st.text_area(
                "Additional request",
                value="Generate a balanced ranked preference list.",
                height=86,
                key="user_request",
            )
            generate_clicked = st.button(
                "Generate Recommendations",
                type="primary",
                use_container_width=True,
                key="generate_recommendations",
            )
    else:
        if st.button(
            "☰",
            key="candidate_panel_toggle_open",
            use_container_width=True,
            help="Show Candidate Profile",
        ):
            st.session_state.show_candidate_panel = True
            st.rerun()

        percentile = float(st.session_state.get("percentile", 90.0))
        category = str(st.session_state.get("category", "OPEN"))
        gender = str(st.session_state.get("gender", "Not Specified"))
        branches = list(st.session_state.get("branches", ["Computer Engineering", "Information Technology"]))
        location_text = str(st.session_state.get("location_text", ""))
        locations = [
            city.strip()
            for city in location_text.split(",")
            if city.strip()
        ]
        preferred_university = str(
            st.session_state.get("preferred_university", "Any")
        )
        seat_type = str(st.session_state.get("seat_type", "Any"))
        college_count = int(st.session_state.get("college_count", 30))
        user_request = str(st.session_state.get("user_request", "Generate a balanced ranked preference list."))
        generate_clicked = False


# ==========================================================
# Knowledge-base panel
# ==========================================================

with right_column:
    if show_right:
        with st.container(border=True):
            if st.button(
                "Knowledge Base  ☰",
                key="knowledge_panel_toggle_close",
                use_container_width=True,
                help="Hide Knowledge Base",
            ):
                st.session_state.show_knowledge_panel = False
                st.rerun()

            st.markdown('<div class="side-heading">📚 Knowledge Base</div>', unsafe_allow_html=True)
            uploaded_files = st.file_uploader(
                "Upload documents",
                accept_multiple_files=True,
                type=["pdf", "docx", "txt", "md"],
                key="knowledge_files",
            )
            document_type = st.selectbox(
                "Document type",
                ["cutoff", "general"],
                key="document_type",
            )
            upload_clicked = st.button(
                "⇧  Upload and Index",
                use_container_width=True,
                key="upload_and_index",
            )
            st.markdown(
                '<div class="side-note">Supported files: PDF, DOCX, TXT and Markdown.<br>'
                'The existing backend upload and indexing API remains unchanged.</div>',
                unsafe_allow_html=True,
            )

            upload_result = st.session_state.get("upload_result")
            if upload_result:
                collection_chunks = upload_result.get(
                    "collection_chunks",
                    upload_result.get("total_chunks_indexed", "—"),
                )
                successful_items = [
                    item for item in upload_result.get("results", [])
                    if isinstance(item, dict) and item.get("status") == "success"
                ]
                st.success(f"Knowledge base ready · Chunks: {collection_chunks}")

                st.markdown("**Indexed Documents**")
                if successful_items:
                    for item in successful_items[:8]:
                        file_name = (
                            item.get("file_name")
                            or item.get("filename")
                            or item.get("source_file")
                            or "Indexed document"
                        )
                        st.caption(f"📄 {file_name}  ·  ✅")
                else:
                    st.caption("Documents were indexed successfully.")

                st.markdown("**Knowledge Base Stats**")
                stat_left, stat_right = st.columns(2)
                with stat_left:
                    st.metric("Documents", len(successful_items) or "—")
                with stat_right:
                    st.metric("Chunks", collection_chunks)
                st.caption("Status: Ready")
    else:
        if st.button(
            "☰",
            key="knowledge_panel_toggle_open",
            use_container_width=True,
            help="Show Knowledge Base",
        ):
            st.session_state.show_knowledge_panel = True
            st.rerun()

        uploaded_files = []
        document_type = "cutoff"
        upload_clicked = False


# ==========================================================
# API actions
# ==========================================================

if upload_clicked:
    if not uploaded_files:
        with right_column:
            st.warning("Select at least one file before uploading.")
    else:
        try:
            multipart_files = [
                ("files", (uploaded.name, uploaded.getvalue(), uploaded.type))
                for uploaded in uploaded_files
            ]
            with right_column, st.spinner("Uploading and indexing documents..."):
                upload_response = requests.post(
                    f"{API_URL}/upload",
                    files=multipart_files,
                    data={"document_type": document_type},
                    timeout=1800,
                )
                if not upload_response.ok:
                    raise RuntimeError(_error_message(upload_response))
                st.session_state.upload_result = upload_response.json()
                st.success("Documents indexed successfully.")
                st.rerun()
        except (requests.RequestException, RuntimeError) as exc:
            with right_column:
                st.error(f"Upload failed: {exc}")

if generate_clicked:
    counselling_body = {
        "student_profile": {
            "percentile": percentile,
            "category": category,
            "gender": None if gender == "Not Specified" else gender,
            "preferred_branches": _without_any(branches),
            "preferred_locations": locations,
            "preferred_university": (
                None
                if preferred_university == "Any"
                else preferred_university
            ),
            "seat_type": None if seat_type == "Any" else seat_type,
            "college_count": college_count,
        },
        "user_request": user_request,
    }

    try:
        with center_column, st.spinner("Counsellor Agent is analysing retrieved evidence..."):
            st.session_state.result = _post_json("/counsel", counselling_body, 1200)
            st.session_state.chat_history = []
            _append_chat_message(
                "assistant",
                "Your recommendation list is ready. You can ask me to add or remove a college, "
                "explain admission risk, compare branches, or suggest better alternatives.",
            )
            st.rerun()
    except (requests.RequestException, RuntimeError) as exc:
        with center_column:
            st.error(f"Counselling request failed: {exc}")


# ==========================================================
# Main dashboard
# ==========================================================

with center_column:
    with st.container(border=True):
        logo_uri = _logo_data_uri()
        logo_markup = (
            f'<img class="hero-logo" src="{logo_uri}" alt="DnyanAI logo">'
            if logo_uri
            else ""
        )
        st.markdown(
            f"""
<div class="hero">
    {logo_markup}
    <div class="hero-title">🎓 MHT-CET Agentic RAG Counsellor V2</div>
    <div class="hero-subtitle">Upload cutoff evidence, generate recommendations, and refine the result through feedback.</div>
</div>
""",
            unsafe_allow_html=True,
        )

        result = st.session_state.get("result")
        recommendations = result.get("recommendations", []) if isinstance(result, dict) else []
        counts = _zone_counts(recommendations)

        st.markdown(
            '<div class="dashboard-title">📊 Master Output Preview Dashboard</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f"""
<div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-label">Total Colleges</div><div class="kpi-value">{len(recommendations)}</div></div>
    <div class="kpi-card"><div class="kpi-label">Dream</div><div class="kpi-value dream">{counts['Dream']}</div></div>
    <div class="kpi-card"><div class="kpi-label">Target</div><div class="kpi-value target">{counts['Target']}</div></div>
    <div class="kpi-card"><div class="kpi-label">Safer</div><div class="kpi-value safer">{counts['Safer']}</div></div>
</div>
""",
            unsafe_allow_html=True,
        )

        if result and recommendations:
            if result.get("summary"):
                st.caption(result["summary"])
            if result.get("strategy"):
                st.info(result["strategy"])

            st.markdown("**Recommended Colleges (Preview)**")
            table = _prepare_table(recommendations)
            table_height = min(545, max(230, 43 + len(table) * 35))
            st.dataframe(
                table,
                use_container_width=True,
                hide_index=True,
                height=table_height,
            )

            requested_count = result.get("requested_college_count")
            if requested_count is not None:
                st.caption(
                    f"Requested colleges: {requested_count} · Generated colleges: {len(recommendations)}"
                )

            if result.get("evidence_warning"):
                st.warning(result["evidence_warning"])

            download_url = result.get("excel_download_url")
            if download_url:
                try:
                    excel_response = requests.get(f"{API_URL}{download_url}", timeout=120)
                    if not excel_response.ok:
                        raise RuntimeError(_error_message(excel_response))
                    st.download_button(
                        "Download Excel",
                        data=excel_response.content,
                        file_name=download_url.rsplit("/", 1)[-1],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except (requests.RequestException, RuntimeError) as exc:
                    st.warning(f"Excel download is currently unavailable: {exc}")
        else:
            st.markdown(
                """
<div class="empty-state">
    <div style="font-size: 2rem;">🎯</div>
    <div class="empty-title">Your recommendation workspace is ready</div>
    <div>Complete the candidate profile, upload cutoff evidence, and generate a ranked CAP preference list.</div>
</div>
""",
                unsafe_allow_html=True,
            )

        st.markdown(
            '<div class="chat-panel"><div class="chat-heading">💬 AI Counsellor Chat</div>'
            '<div class="quick-action-note">Ask naturally: “Add VJTI IT”, “Show better branches”, '
            '“Suggest safer CS colleges”, “Remove college 5”, or “Yes, add it anyway”.</div></div>',
            unsafe_allow_html=True,
        )

        if not st.session_state.chat_history:
            with st.chat_message("assistant", avatar="🎓"):
                st.markdown(
                    "Generate recommendations first. After that, I can explain cutoffs, "
                    "suggest alternatives, and update your preference list through conversation."
                )
        else:
            for message in st.session_state.chat_history:
                avatar = "🎓" if message.get("role") == "assistant" else "👤"
                with st.chat_message(message.get("role", "assistant"), avatar=avatar):
                    st.markdown(str(message.get("content", "")))

        session_id = result.get("session_id") if isinstance(result, dict) else None

        if isinstance(result, dict):
            _render_feedback_actions(result, session_id)

        chat_prompt = st.chat_input(
            "Ask your counsellor or confirm a pending choice...",
            disabled=not bool(session_id),
            key=f"feedback_chat_{st.session_state.chat_input_version}",
        )

        if chat_prompt:
            _send_feedback_action(
                session_id,
                chat_prompt.strip(),
                chat_prompt.strip(),
            )

        if st.session_state.chat_history:
            clear_left, clear_right = st.columns([1, 5])
            with clear_left:
                if st.button("Clear Chat", key="clear_counsellor_chat"):
                    st.session_state.chat_history = []
                    st.session_state.chat_input_version += 1
                    st.rerun()